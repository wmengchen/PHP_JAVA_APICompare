# -*- coding:utf-8 -*-
# @Time    : 2025/09/16
# @Author  : mengchen.wang
# @File    : PHP&Java_compare.py
# 功能说明 : 支持多账号循环跑接口对比
#           - 用例信息从 sheet【接口信息】读取
#           - 账号信息从 sheet【account_id】读取
#           - 支持限制最大账号数(MAX_ACCOUNTS)
#           - 支持只跑指定账号(ACCOUNT_FILTER)
#           - Header 自动加上 X-Internal-Call: DELAY
# ****************************

import os
import json
import requests
import pytest
import allure
from deepdiff import DeepDiff
from openpyxl import load_workbook

# ===============================
# 配置区
# ===============================
EXCEL_FILE = r"F:\26_接口对比测试\customer\api_cases.xlsx"   # Excel 用例文件
REPORT_PATH = r"C:\report"                                   # allure 报告路径
ENV = "test"  # 可选: "test" / "prod"

# ✅ 最大账号数 (None 表示不限制)
MAX_ACCOUNTS = 1

# ✅ 指定要跑的账号（优先级高于 MAX_ACCOUNTS）
ACCOUNT_FILTER = ["492"]  #  [] 表示指定账号信息
# ACCOUNT_FILTER = []  # 为空列表 [] 表示不指定

# PHP/Java 域名
PHP_DOMAIN = "https://members-staging4.helium-staging.com"
JAVA_DOMAIN = "http://h10api-test.pacvue.com/customer-api"

# 超级token 前缀（根据环境拼接）
TOKEN_PREFIX = {
    "test": "Bearer m22d3bLeUztqGfWtwfAxFw5sUrdqVgsT02A_hk0S7nv0NElyM493_mXKDLsy-vIf_",
    "prod": "Bearer n8_3T-b3udkMDxfCvVF1hEFXY8xUcS0Y788yIIIyQmIZtJOy4BRpSJ43-7ysmxw2_"
}

# ===============================
# 读取账号信息
# ===============================
def read_account_ids(file_path, sheet_name="account_id"):
    """
    读取账号信息 sheet，返回 account_id 列表
    """
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]

    account_ids = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            account_ids.append(str(row[0]).strip())

    # 如果指定了 ACCOUNT_FILTER，则只取这些账号
    if ACCOUNT_FILTER:
        account_ids = [acc for acc in account_ids if acc in ACCOUNT_FILTER]
    # 否则按 MAX_ACCOUNTS 限制
    elif MAX_ACCOUNTS is not None:
        account_ids = account_ids[:MAX_ACCOUNTS]

    return account_ids

# ===============================
# 读取接口信息
# ===============================
def read_excel_cases(file_path, sheet_name="接口信息"):
    """
    读取接口信息，并为每个 account_id 展开用例
    """
    account_ids = read_account_ids(file_path)
    if not account_ids:
        raise ValueError("未在账号信息 sheet 中找到有效的 Account_id")

    wb = load_workbook(file_path)
    sheet = wb[sheet_name]

    apis = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 跳过表头
        if not row or not row[0]:
            continue

        name = str(row[0]).strip()
        method = str(row[1]).strip().upper() if row[1] else "GET"

        # 请求头
        base_headers = {}
        if row[2]:
            try:
                base_headers = json.loads(row[2])
            except Exception:
                print(f"⚠️ 警告: {name} 的请求头解析失败，将忽略。")

        path = str(row[3]).strip() if row[3] else ""

        # 公共参数
        common_params = {}
        if row[4]:
            try:
                common_params = json.loads(row[4])
            except Exception:
                print(f"⚠️ 警告: {name} 的公共参数解析失败，将忽略。")

        # PHP 参数
        php_params = {}
        if row[5]:
            try:
                php_params = json.loads(row[5])
            except Exception:
                print(f"⚠️ 警告: {name} 的 PHP 参数解析失败，将忽略。")

        # Java 参数
        java_params = {}
        if row[6]:
            try:
                java_params = json.loads(row[6])
            except Exception:
                print(f"⚠️ 警告: {name} 的 Java 参数解析失败，将忽略。")

        # 忽略字段
        ignore_fields = [f.strip() for f in str(row[7]).split(",")] if row[7] else []

        # 判断使用公共参数还是独立参数
        if common_params:
            php_final_params = common_params
            java_final_params = common_params
        else:
            php_final_params = php_params
            java_final_params = java_params

        # 🔑 展开：对每个 account_id 生成一个 case
        for account_id in account_ids:
            headers = dict(base_headers)  # 复制一份
            headers["Authorization"] = TOKEN_PREFIX[ENV] + account_id
            headers["X-Internal-Call"] = "DELAY"   # ✅ 自动加上该参数

            apis.append({
                "name": f"{name}_acc{account_id}",  # 区分不同账号的 case
                "method": method,
                "headers": headers,
                "path": path,
                "php_params": php_final_params,
                "java_params": java_final_params,
                "ignore_fields": ignore_fields,
                "account_id": account_id
            })

    return apis

# ===============================
# 通用请求
# ===============================
def send_request(method, url, params, headers=None):
    """ 通用请求方法，支持 headers """
    try:
        method = method.upper()
        if method == "GET":
            response = requests.get(url, params=params, headers=headers, timeout=10)
        elif method == "POST":
            response = requests.post(url, json=params, headers=headers, timeout=10)
        elif method == "PUT":
            response = requests.put(url, json=params, headers=headers, timeout=10)
        elif method == "DELETE":
            response = requests.delete(url, json=params, headers=headers, timeout=10)
        else:
            return {"error": f"不支持的请求方式: {method}"}
        return response.json()
    except Exception as e:
        return {"error": str(e)}

# ===============================
# 删除忽略字段
# ===============================
def remove_fields(data, ignore_fields):
    """
    递归删除字段，支持路径通配符*，如 results.*.createdAt，也支持数组下标
    """
    def _remove(d, keys):
        if not d or not keys:
            return
        key = keys[0]
        if key == "*":  # 支持通配符 *
            if isinstance(d, dict):
                for subkey in list(d.keys()):
                    _remove(d[subkey], keys[1:])
            elif isinstance(d, list):
                for item in d:
                    _remove(item, keys[1:])
        else:
            if isinstance(d, list):
                try:
                    idx = int(key)
                    if idx < len(d):
                        _remove(d[idx], keys[1:])
                except:
                    for item in d:
                        _remove(item, keys)
            elif isinstance(d, dict):
                if len(keys) == 1:
                    d.pop(key, None)
                elif key in d:
                    _remove(d[key], keys[1:])
    data_copy = json.loads(json.dumps(data))  # 深拷贝，防止污染原始数据
    for field in ignore_fields:
        _remove(data_copy, field.split('.'))
    return data_copy

# ===============================
# pytest + allure 测试用例
# ===============================
@allure.epic("PHP&Java接口一致性对比测试")
class TestAPICompare:

    @pytest.mark.parametrize("api", read_excel_cases(EXCEL_FILE))
    def test_api_compare(self, api):
        test_name = f"test_{api['name']}"
        allure.dynamic.title(test_name)

        name = api["name"]
        method = api["method"]
        path = api["path"]
        php_params = api["php_params"]
        java_params = api["java_params"]
        ignore_fields = api["ignore_fields"]
        headers = api["headers"]

        php_url = PHP_DOMAIN + path
        java_url = JAVA_DOMAIN + path

        with allure.step(f"请求 PHP 接口: {method} {php_url}"):
            php_resp = send_request(method, php_url, php_params, headers)
            allure.attach(json.dumps(php_resp, indent=2, ensure_ascii=False),
                        name="PHP返回", attachment_type=allure.attachment_type.JSON)

        with allure.step(f"请求 Java 接口: {method} {java_url}"):
            java_resp = send_request(method, java_url, java_params, headers)
            allure.attach(json.dumps(java_resp, indent=2, ensure_ascii=False),
                        name="Java返回", attachment_type=allure.attachment_type.JSON)

        # 对比返回内容
        with allure.step("对比返回内容"):
            php_filtered = remove_fields(php_resp, ignore_fields)
            java_filtered = remove_fields(java_resp, ignore_fields)
            diff = DeepDiff(php_filtered, java_filtered, ignore_order=True)
            if diff:
                allure.attach(
                    diff.to_json(indent=2, ensure_ascii=False),
                    name="差异详情",
                    attachment_type=allure.attachment_type.JSON
                )
                pytest.fail(f"接口 {name} 返回结果不一致")

# ===============================
# main入口
# ===============================
if __name__ == "__main__":
    result_dir = f"{REPORT_PATH}\\result"
    if os.path.exists(result_dir):
        import shutil
        shutil.rmtree(result_dir)
    os.makedirs(result_dir, exist_ok=True)
    
    pytest.main([
        __file__,
        f"--alluredir={result_dir}",
        "-q", "-s"
    ])

    os.system(f"allure generate {result_dir} -o {REPORT_PATH}\\html --clean")
    print(f"测试完成，报告已生成: {REPORT_PATH}\\html\\index.html")
    os.system(f"allure open {REPORT_PATH}\\html")
